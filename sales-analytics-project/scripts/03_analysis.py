"""
03_analysis.py
--------------
Runs summary analysis on the cleaned sales data and produces:
 - Static charts (PNG) for the README / portfolio
 - An Excel workbook with pivot-style summary tables and formulas
   (the PowerPivot-equivalent deliverable for this dataset)

Run: python 03_analysis.py
Input:  ../data/cleaned_sales_data.csv
Output: ../charts/*.png, ../Sales_Summary.xlsx
"""

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

plt.rcParams["font.family"] = "DejaVu Sans"
COLORS = ["#2E5A88", "#4C8BF5", "#7FB2F0", "#F2A65A", "#E85D75"]

df = pd.read_csv("../data/cleaned_sales_data.csv", parse_dates=["OrderDate"])
df["Month"] = df["OrderDate"].dt.to_period("M").astype(str)

# ---------- Chart 1: Revenue by Region ----------
region_rev = df.groupby("Region")["Revenue"].sum().sort_values(ascending=False)

fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.bar(region_rev.index, region_rev.values, color=COLORS)
ax.set_title("Total Revenue by Region", fontsize=14, fontweight="bold")
ax.set_ylabel("Revenue (ZAR)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"R{x:,.0f}"))
plt.xticks(rotation=20, ha="right")
for b in bars:
    ax.text(b.get_x() + b.get_width()/2, b.get_height(), f"R{b.get_height():,.0f}",
            ha="center", va="bottom", fontsize=9)
plt.tight_layout()
plt.savefig("../charts/revenue_by_region.png", dpi=150)
plt.close()

# ---------- Chart 2: Monthly Revenue Trend ----------
monthly_rev = df.groupby("Month")["Revenue"].sum().sort_index()

fig, ax = plt.subplots(figsize=(9, 5))
ax.plot(monthly_rev.index, monthly_rev.values, marker="o", color="#2E5A88", linewidth=2)
ax.fill_between(range(len(monthly_rev)), monthly_rev.values, color="#4C8BF5", alpha=0.15)
ax.set_title("Monthly Revenue Trend (2025)", fontsize=14, fontweight="bold")
ax.set_ylabel("Revenue (ZAR)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"R{x:,.0f}"))
plt.xticks(rotation=45, ha="right")
plt.tight_layout()
plt.savefig("../charts/monthly_trend.png", dpi=150)
plt.close()

# ---------- Chart 3: Top Products ----------
top_products = df.groupby("Product")["Revenue"].sum().sort_values(ascending=False).head(6)

fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.barh(top_products.index[::-1], top_products.values[::-1], color=COLORS[1])
ax.set_title("Top Products by Revenue", fontsize=14, fontweight="bold")
ax.set_xlabel("Revenue (ZAR)")
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"R{x:,.0f}"))
plt.tight_layout()
plt.savefig("../charts/top_products.png", dpi=150)
plt.close()

# ---------- Chart 4: Category share ----------
cat_rev = df.groupby("Category")["Revenue"].sum().sort_values(ascending=False)
fig, ax = plt.subplots(figsize=(6.5, 6.5))
ax.pie(cat_rev.values, labels=cat_rev.index, autopct="%1.1f%%",
       colors=COLORS, startangle=90, wedgeprops={"edgecolor": "white", "linewidth": 1})
ax.set_title("Revenue Share by Category", fontsize=14, fontweight="bold")
plt.tight_layout()
plt.savefig("../charts/category_share.png", dpi=150)
plt.close()

print("Charts saved to ../charts/")

# =========================================================
# Excel workbook: pivot-style summary (PowerPivot-equivalent)
# =========================================================
wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="2E5A88")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=13, color="2E5A88")

def style_header(ws, row=1, cols=None):
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

def autosize(ws):
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 3

# ---- Sheet 1: Raw-ish cleaned data (reference) ----
ws1 = wb.active
ws1.title = "Cleaned Data"
ws1.append(["OrderID", "OrderDate", "Region", "Product", "Category",
            "Quantity", "UnitPrice", "SalesRep", "Revenue"])
for _, r in df.iterrows():
    ws1.append([r["OrderID"], r["OrderDate"].strftime("%Y-%m-%d"), r["Region"],
                r["Product"], r["Category"], int(r["Quantity"]),
                round(r["UnitPrice"], 2), r["SalesRep"], round(r["Revenue"], 2)])
style_header(ws1)
autosize(ws1)

# ---- Sheet 2: Pivot - Revenue by Region & Category ----
ws2 = wb.create_sheet("Pivot - Region x Category")
pivot = pd.pivot_table(df, values="Revenue", index="Region", columns="Category",
                        aggfunc="sum", fill_value=0, margins=True, margins_name="Grand Total")
ws2.append(["Region"] + list(pivot.columns))
for idx, row in pivot.iterrows():
    ws2.append([idx] + [round(v, 2) for v in row])
style_header(ws2)
autosize(ws2)

bar = BarChart()
bar.title = "Revenue by Region and Category"
bar.y_axis.title = "Revenue (ZAR)"
bar.x_axis.title = "Region"
data_ref = Reference(ws2, min_col=2, max_col=pivot.shape[1], min_row=1,
                      max_row=pivot.shape[0])  # exclude grand total row/col for cleaner chart
cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=pivot.shape[0])
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.width, bar.height = 22, 12
ws2.add_chart(bar, f"A{pivot.shape[0] + 4}")

# ---- Sheet 3: Pivot - Monthly Trend ----
ws3 = wb.create_sheet("Pivot - Monthly Trend")
monthly = df.groupby("Month").agg(
    Orders=("OrderID", "count"),
    Units=("Quantity", "sum"),
    Revenue=("Revenue", "sum")
).reset_index()
ws3.append(["Month", "Orders", "Units Sold", "Revenue"])
for _, r in monthly.iterrows():
    ws3.append([r["Month"], int(r["Orders"]), int(r["Units"]), round(r["Revenue"], 2)])
style_header(ws3)
autosize(ws3)

line = LineChart()
line.title = "Monthly Revenue Trend"
line.y_axis.title = "Revenue (ZAR)"
line.x_axis.title = "Month"
data_ref = Reference(ws3, min_col=4, min_row=1, max_row=monthly.shape[0] + 1)
cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=monthly.shape[0] + 1)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.width, line.height = 22, 12
ws3.add_chart(line, "F2")

# ---- Sheet 4: KPI Summary (with live Excel formulas) ----
ws4 = wb.create_sheet("KPI Summary", 0)  # make it the first visible sheet
ws4["A1"] = "Sales Performance – KPI Summary"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:B1")

n_rows = len(df)
data_range_rev = f"'Cleaned Data'!I2:I{n_rows + 1}"
data_range_qty = f"'Cleaned Data'!F2:F{n_rows + 1}"

kpis = [
    ("Total Revenue", f"=SUM({data_range_rev})"),
    ("Total Units Sold", f"=SUM({data_range_qty})"),
    ("Total Orders", f"=COUNTA('Cleaned Data'!A2:A{n_rows + 1})"),
    ("Average Order Value", f"=SUM({data_range_rev})/COUNTA('Cleaned Data'!A2:A{n_rows + 1})"),
    ("Top Region (by Revenue)", "Gauteng (see Pivot sheet)"),
]
ws4.append([])
ws4.append(["Metric", "Value"])
style_header(ws4, row=3)
for label, formula in kpis:
    ws4.append([label, formula])
for row in ws4.iter_rows(min_row=4, max_row=3 + len(kpis), min_col=2, max_col=2):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.number_format = '"R"#,##0.00'
autosize(ws4)

wb.save("../Sales_Summary.xlsx")
print("Excel workbook saved -> ../Sales_Summary.xlsx")

# ---------- Console summary (also used to populate README stats) ----------
print("\n--- Quick Summary ---")
print(f"Total Revenue: R{df['Revenue'].sum():,.2f}")
print(f"Total Orders: {len(df)}")
print(f"Top Region: {region_rev.idxmax()} (R{region_rev.max():,.2f})")
print(f"Top Product: {top_products.idxmax()} (R{top_products.max():,.2f})")
